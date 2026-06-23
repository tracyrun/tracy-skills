"""
可研报告投资估算表 Excel 生成脚本 — 含公式链接版（修订版）
依据：设计院2026.05.27总图t8版 + 建构筑物一览表
项目总资金：65,656万元
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

def generate_investment_table(output_dir=None):
    wb = openpyxl.Workbook()

    # ========== Styles ==========
    header_font_w = Font(name='微软雅黑', bold=True, size=9, color='FFFFFF')
    title_font = Font(name='微软雅黑', bold=True, size=14)
    subtitle_font = Font(name='微软雅黑', bold=True, size=11)
    normal_font = Font(name='微软雅黑', size=9)
    bold_font = Font(name='微软雅黑', bold=True, size=9)
    total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    section_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # ========== Sheet 1: 投资估算表 ==========
    ws = wb.active
    ws.title = "投资估算表"

    col_widths = [5.5, 30, 12, 11, 11, 11, 13, 8, 11, 11, 20, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = '德阳市大宗货物（旌北新区）物流基地项目（黄许一期）— 投资估算表'
    ws['A1'].font = title_font; ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:L2')
    ws['A2'] = '金额单位：万元  |  修订版：依据设计院2026.05.27总图t8版  |  编制日期：2026年6月'
    ws['A2'].font = Font(name='微软雅黑', size=8, color='666666')
    ws['A2'].alignment = Alignment(horizontal='right', vertical='center')

    # Headers (rows 4-5, 2-level)
    def make_header():
        ws.merge_cells('A4:A5'); ws.merge_cells('B4:B5')
        ws.merge_cells('C4:G4'); ws.merge_cells('H4:J4')
        ws.merge_cells('K4:K5'); ws.merge_cells('L4:L5')

        labels_r4 = {1: '序号', 2: '工程或费用名称', 3: '估算价值（万元）', 8: '技术经济指标', 11: '备注', 12: '比例\n（%）'}
        for col, label in labels_r4.items():
            c = ws.cell(row=4, column=col, value=label)
            c.font = header_font_w; c.alignment = center_align; c.fill = header_fill; c.border = thin_border

        sub_h = ['建筑工程费', '安装工程费', '设备购置费', '其他费用', '合计']
        for i, sh in enumerate(sub_h):
            c = ws.cell(row=5, column=3+i, value=sh)
            c.font = header_font_w; c.alignment = center_align; c.fill = header_fill; c.border = thin_border

        sub_t = ['单位', '数量', '单位价值\n（元/单位）']
        for i, sh in enumerate(sub_t):
            c = ws.cell(row=5, column=8+i, value=sh)
            c.font = header_font_w; c.alignment = center_align; c.fill = header_fill; c.border = thin_border

        for col in [1, 2, 11, 12]:
            ws.cell(row=5, column=col).fill = header_fill
            ws.cell(row=5, column=col).border = thin_border
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 35

    make_header()

    # ========== Data structure ==========
    entries = []
    row_num = 6

    def add_entry(seq, name, row_type, bld=None, inst=None, equip=None, other=None,
                  unit=None, qty=None, unit_val=None, cost_col=None, remark=None, ratio=None, bold=False):
        nonlocal row_num
        entries.append({
            'row': row_num, 'seq': seq, 'name': name, 'row_type': row_type,
            'bld': bld, 'inst': inst, 'equip': equip, 'other': other,
            'unit': unit, 'qty': qty, 'unit_val': unit_val, 'cost_col': cost_col,
            'remark': remark, 'ratio': ratio, 'bold': bold
        })
        row_num += 1

    # ===== 一、工程费用 =====
    add_entry('一', '工程费用', 'section', ratio=72.9, bold=True)

    # 1. 建筑工程 (15 items)
    add_entry('1', '建筑工程', 'subtotal', ratio=49.5, bold=True)
    add_entry('1.1', '大宗物流仓 1#', 'item', unit='㎡', qty=23040, unit_val=1550,
              cost_col=3, remark='2F钢结构，含坡道+货梯土建，最大单体')
    add_entry('1.2', '大宗物流仓 2#', 'item', unit='㎡', qty=23040, unit_val=1530,
              cost_col=3, remark='2F钢结构，最大单体')
    add_entry('1.3', '大宗物流仓 3#', 'item', unit='㎡', qty=8160, unit_val=1580,
              cost_col=3, remark='2F钢结构，小单体单价略高')
    add_entry('1.4', '大宗物流仓 4#', 'item', unit='㎡', qty=11952, unit_val=1550,
              cost_col=3, remark='2F钢结构')
    add_entry('1.5', '大宗物流仓 5#', 'item', unit='㎡', qty=17424, unit_val=1530,
              cost_col=3, remark='2F钢结构')
    add_entry('1.6', '大宗物流仓 6#', 'item', unit='㎡', qty=16128, unit_val=1530,
              cost_col=3, remark='2F钢结构')
    add_entry('1.7', '大宗物流仓 7#', 'item', unit='㎡', qty=16128, unit_val=1530,
              cost_col=3, remark='2F钢结构')
    add_entry('1.8', '大宗物流仓 8#', 'item', unit='㎡', qty=13440, unit_val=1530,
              cost_col=3, remark='2F钢结构')
    add_entry('1.9', '大宗物流仓 11#', 'item', unit='㎡', qty=6912, unit_val=1580,
              cost_col=3, remark='2F钢结构，小单体')
    add_entry('1.10', '冷库 9#', 'item', unit='㎡', qty=13440, unit_val=6800,
              cost_col=3, remark='2F，NH₃/CO₂复叠制冷，150mm PIR保温')
    add_entry('1.11', '中转仓 10#', 'item', unit='㎡', qty=6912, unit_val=1620,
              cost_col=3, remark='2F钢结构，含双侧月台')
    add_entry('1.12', '机修间', 'item', unit='㎡', qty=1008, unit_val=2200,
              cost_col=3, remark='1F钢结构')
    add_entry('1.13', '设备间', 'item', unit='㎡', qty=1008, unit_val=2200,
              cost_col=3, remark='1F钢结构')
    add_entry('1.14', '垃圾间', 'item', unit='㎡', qty=227, unit_val=1500,
              cost_col=3, remark='1F钢结构')
    add_entry('1.15', '防火涂料增项（一级耐火）', 'item', unit='㎡', qty=158819, unit_val=45,
              cost_col=3, remark='厚涂型防火涂料，钢结构一级耐火')

    # 2. 安装工程 (5 items)
    add_entry('2', '安装工程', 'subtotal', ratio=7.7, bold=True)
    add_entry('2.1', '给排水工程', 'item', unit='㎡', qty=158819, unit_val=35,
              cost_col=4, remark='含雨污水管网、给水系统')
    add_entry('2.2', '供配电工程', 'item', unit='㎡', qty=158819, unit_val=88,
              cost_col=4, remark='含高低压配电、照明，冷库用电负荷增加')
    add_entry('2.3', '暖通工程', 'item', unit='㎡', qty=158819, unit_val=32,
              cost_col=4, remark='通风、排烟、冷库制冷管道')
    add_entry('2.4', '消防工程', 'item', unit='㎡', qty=158819, unit_val=125,
              cost_col=4, remark='一级耐火增项：含消火栓、ESFR喷淋、报警')
    add_entry('2.5', '智能化工程', 'item', unit='㎡', qty=158819, unit_val=42,
              cost_col=4, remark='含监控、网络、WMS基础设施')

    # 3. 设备购置费 (4 items)
    add_entry('3', '设备购置费', 'subtotal', ratio=10.5, bold=True)
    add_entry('3.1', '物流设备', 'item', unit='项', qty=1, unit_val=36060000,
              cost_col=5, remark='叉车+货架+货梯+托盘+地磅等（详见5.2.2）')
    add_entry('3.2', '冷链设备', 'item', unit='项', qty=1, unit_val=28320000,
              cost_col=5, remark='压缩机组+冷风机+冷凝器+冷库门（详见5.2.3）')
    add_entry('3.3', '信息化设备', 'item', unit='项', qty=1, unit_val=6560000,
              cost_col=5, remark='服务器+网络+监控+大屏+UPS（详见5.2.4）')
    add_entry('3.4', '充电桩', 'item', unit='台', qty=48, unit_val=4000,
              cost_col=5, remark='预留15%充电车位')

    # 4. 室外配套工程 (9 items)
    add_entry('4', '室外配套工程', 'subtotal', ratio=5.1, bold=True)
    add_entry('4.1', '道路及硬化', 'item', unit='㎡', qty=72000, unit_val=285,
              cost_col=6, remark='含主干道(沥青)+次干道+支路(混凝土)')
    add_entry('4.2', '绿化工程', 'item', unit='㎡', qty=19650, unit_val=82,
              cost_col=6, remark='含行道树、绿地、海绵设施')
    add_entry('4.3', '货车停车场', 'item', unit='㎡', qty=10100, unit_val=320,
              cost_col=6, remark='C35重载混凝土')
    add_entry('4.4', '给排水外网', 'item', unit='项', qty=1, unit_val=5200000,
              cost_col=6, remark='接入市政管网，含消防管网')
    add_entry('4.5', '供配电外线', 'item', unit='项', qty=1, unit_val=3500000,
              cost_col=6, remark='10kV外线+箱变')
    add_entry('4.6', '围墙及大门', 'item', unit='m', qty=1800, unit_val=1200,
              cost_col=6, remark='2.5m砖砌+铁艺，4座电动门')
    add_entry('4.7', '海绵城市增量', 'item', unit='项', qty=1, unit_val=1800000,
              cost_col=6, remark='雨水花园+下凹绿地+透水铺装+调蓄池')
    add_entry('4.8', '绿色建筑增量', 'item', unit='项', qty=1, unit_val=1200000,
              cost_col=6, remark='一星级绿色建筑增量成本')
    add_entry('4.9', '室外综合管沟', 'item', unit='m', qty=2800, unit_val=350,
              cost_col=6, remark='给水/消防/电力/通信共沟')

    # ===== 二、工程建设其他费 =====
    add_entry('二', '工程建设其他费', 'section', ratio=14.1, bold=True)
    add_entry('5', '土地取得费', 'item', other=True, unit='亩', qty=295, unit_val=180000,
              remark='含土地出让金及税费，按180,000元/亩综合价', ratio=8.1)
    add_entry('6', '建设单位管理费', 'item', other=True, bld=530,
              remark='财建〔2016〕504号阶梯费率', ratio=0.8)
    add_entry('7', '勘察设计费', 'item', other=True, bld=1193,
              remark='工程费用×2.5%', ratio=1.8)
    add_entry('8', '工程监理费', 'item', other=True, bld=716,
              remark='发改价格〔2007〕670号插入法', ratio=1.1)
    add_entry('9', '全过程造价咨询', 'item', other=True, bld=140,
              remark='含概算编制、全过程造价控制、结算审核', ratio=0.2)
    add_entry('10', '招标代理费', 'item', other=True, bld=95,
              remark='6个标段招标代理', ratio=0.1)
    add_entry('11', '前期工作咨询费', 'item', other=True, bld=150,
              remark='可研编制、环评、水保、节能评估', ratio=0.2)
    add_entry('12', '专项评价费', 'item', other=True, bld=160,
              remark='地震安评、地灾评估、职业病危害预评价', ratio=0.2)
    add_entry('13', '工程保险费', 'item', other=True, bld=143,
              remark='工程费用×0.3%', ratio=0.2)
    add_entry('14', '城市基础设施配套费', 'item', other=True, bld=318,
              remark='按当地标准（158,819㎡×20元/㎡）', ratio=0.5)
    add_entry('15', 'BIM技术应用费', 'item', other=True, bld=127,
              remark='设计+BIM+运维基础模型', ratio=0.2)
    add_entry('16', '临时设施费', 'item', other=True, bld=238,
              remark='工程费用×0.5%，施工临时水电道路围挡', ratio=0.4)
    add_entry('17', '检验检测费', 'item', other=True, bld=95,
              remark='地基检测+钢结构探伤+消防检测+防雷检测', ratio=0.1)

    # ===== 三、预备费 =====
    add_entry('三', '预备费', 'section', ratio=7.0, bold=True)
    add_entry('18', '基本预备费', 'item', other=True,
              remark='(工程费用+工程建设其他费)×8%', ratio=7.0)

    # ===== 四、建设期利息 =====
    add_entry('四', '建设期利息', 'section', ratio=2.6, bold=True)
    add_entry('19', '建设期融资利息', 'item', other=True, bld=1680,
              remark='按融资30,000万元×2.8%×2年（平均）', ratio=2.6)

    # ===== 五、项目总投资 =====
    add_entry('五', '项目总投资', 'grand_total', ratio=96.2, bold=True)

    # ===== 铺底流动资金 =====
    add_entry('', '铺底流动资金', 'item', other=True, bld=2500,
              remark='按运营初期3个月运营成本估算', ratio=3.8)

    # ===== 六、项目总资金 =====
    add_entry('六', '项目总资金', 'grand_total', ratio=100.0, bold=True)

    # ===== Collect section/subtotal ranges =====
    section_ranges = {}
    current_section = None
    for e in entries:
        if e['row_type'] == 'section':
            current_section = e['name']
            section_ranges[current_section] = {'start': e['row'], 'end': e['row'], 'subtotals': []}
        elif current_section:
            section_ranges[current_section]['end'] = e['row']
            if e['row_type'] == 'subtotal':
                section_ranges[current_section]['subtotals'].append(e['row'])

    subtotal_ranges = []
    current_subtotal = None
    current_subtotal_start = None
    for e in entries:
        if e['row_type'] in ('subtotal',):
            if current_subtotal is not None:
                subtotal_ranges.append((current_subtotal, current_subtotal_start, e['row'] - 1))
            current_subtotal = e['row']
            current_subtotal_start = e['row'] + 1
        elif e['row_type'] == 'section':
            if current_subtotal is not None:
                subtotal_ranges.append((current_subtotal, current_subtotal_start, e['row'] - 1))
            current_subtotal = None; current_subtotal_start = None
        elif e['row_type'] == 'grand_total':
            if current_subtotal is not None:
                subtotal_ranges.append((current_subtotal, current_subtotal_start, e['row'] - 1))
            current_subtotal = None

    # ===== Write cells =====
    for e in entries:
        row = e['row']
        is_section = e['row_type'] == 'section'
        is_subtotal = e['row_type'] == 'subtotal'
        is_grand = e['row_type'] == 'grand_total'
        is_bold = e['bold'] or is_section or is_subtotal or is_grand
        fill = section_fill if is_section else (total_fill if (is_subtotal or is_grand) else None)

        # A: 序号
        c = ws.cell(row=row, column=1, value=e['seq'] if e['seq'] != '' else None)
        c.font = bold_font if is_bold else normal_font; c.alignment = center_align; c.border = thin_border

        # B: 名称
        c = ws.cell(row=row, column=2, value=e['name'])
        c.font = bold_font if is_bold else normal_font; c.alignment = left_align; c.border = thin_border

        has_unit_formula = (e['unit'] and e['qty'] is not None and e['unit_val'] is not None)

        # C-G: 估算价值
        if not (is_section or is_subtotal or is_grand):
            if has_unit_formula:
                formula_g = f'=I{row}*J{row}/10000'
                ws.cell(row=row, column=7).value = formula_g
                cc = e.get('cost_col')
                if cc:
                    ws.cell(row=row, column=cc).value = f'=G{row}'
                elif e.get('other'):
                    ws.cell(row=row, column=6).value = f'=G{row}'
            elif e.get('other'):
                bld_val = e.get('bld')
                if bld_val is not None:
                    ws.cell(row=row, column=6).value = bld_val
                    ws.cell(row=row, column=7).value = f'=F{row}'

        # Style C-G
        for col in range(3, 8):
            c = ws.cell(row=row, column=col)
            c.font = bold_font if is_bold else normal_font
            c.alignment = right_align; c.border = thin_border
            if c.value and isinstance(c.value, (int, float)):
                c.number_format = '#,##0.00'
            elif c.value and isinstance(c.value, str) and c.value.startswith('='):
                c.number_format = '#,##0.00'

        # H-J: 技术经济指标
        for col, val in [(8, e.get('unit')), (9, e.get('qty')), (10, e.get('unit_val'))]:
            if val:
                c = ws.cell(row=row, column=col, value=val)
                c.font = normal_font; c.border = thin_border
                if col in (9, 10) and isinstance(val, (int, float)):
                    c.number_format = '#,##0'
                    c.alignment = right_align
                else:
                    c.alignment = center_align
            else:
                ws.cell(row=row, column=col).border = thin_border

        # K: 备注
        if e.get('remark'):
            c = ws.cell(row=row, column=11, value=e['remark'])
            c.font = normal_font; c.alignment = left_align; c.border = thin_border
        else:
            ws.cell(row=row, column=11).border = thin_border

        # L: 比例
        if e.get('ratio') is not None:
            ratio_val = e['ratio']
            c = ws.cell(row=row, column=12,
                        value=ratio_val/100 if isinstance(ratio_val, (int, float)) else ratio_val)
            c.font = bold_font if is_bold else normal_font
            c.alignment = center_align; c.border = thin_border
            c.number_format = '0.0%'
        else:
            ws.cell(row=row, column=12).border = thin_border

        # Fill rows
        if fill:
            for col in range(1, 13):
                ws.cell(row=row, column=col).fill = fill

    # ===== Subtotal SUM formulas =====
    for sub_row, first_child, last_child in subtotal_ranges:
        if last_child >= first_child:
            formula = f'=SUM(G{first_child}:G{last_child})'
            ws.cell(row=sub_row, column=7).value = formula
            ws.cell(row=sub_row, column=7).number_format = '#,##0.00'
            # Set cost column for subtotal
            for e in entries:
                if e['row'] == sub_row:
                    cc = e.get('cost_col')
                    if cc:
                        ws.cell(row=sub_row, column=cc).value = f'=G{sub_row}'
                    break

    # ===== Section formulas =====
    eng_row = section_ranges['工程费用']['start']
    other_row = section_ranges['工程建设其他费']['start']
    reserve_row = section_ranges['预备费']['start']
    interest_row = section_ranges['建设期利息']['start']

    # 工程费用 section: distribute subtotals to C/D/E/F
    sub_rows_eng = section_ranges['工程费用']['subtotals']
    if len(sub_rows_eng) >= 4:
        ws.cell(row=eng_row, column=3).value = f'=G{sub_rows_eng[0]}'  # 建筑工程
        ws.cell(row=eng_row, column=4).value = f'=G{sub_rows_eng[1]}'  # 安装
        ws.cell(row=eng_row, column=5).value = f'=G{sub_rows_eng[2]}'  # 设备
        ws.cell(row=eng_row, column=6).value = f'=G{sub_rows_eng[3]}'  # 室外配套
        ws.cell(row=eng_row, column=7).value = f'=C{eng_row}+D{eng_row}+E{eng_row}+F{eng_row}'
        for col in range(3, 8):
            ws.cell(row=eng_row, column=col).number_format = '#,##0.00'
            ws.cell(row=eng_row, column=col).font = bold_font
            ws.cell(row=eng_row, column=col).alignment = right_align

    # 工程建设其他费 section: SUM of item rows
    child_start = other_row + 1
    child_end = section_ranges['工程建设其他费']['end']
    data_rows = [e['row'] for e in entries
                 if child_start <= e['row'] <= child_end and e['row_type'] == 'item']
    if data_rows:
        ws.cell(row=other_row, column=7).value = f'=SUM(G{data_rows[0]}:G{data_rows[-1]})'
    ws.cell(row=other_row, column=6).value = f'=G{other_row}'
    ws.cell(row=other_row, column=7).number_format = '#,##0.00'
    ws.cell(row=other_row, column=6).number_format = '#,##0.00'

    # 预备费 section: (工程费用 + 其他费) * 8%
    reserve_formula = f'=ROUND((G{eng_row}+G{other_row})*0.08, 2)'
    ws.cell(row=reserve_row, column=7).value = reserve_formula
    ws.cell(row=reserve_row, column=6).value = f'=G{reserve_row}'
    ws.cell(row=reserve_row, column=7).number_format = '#,##0.00'
    ws.cell(row=reserve_row, column=6).number_format = '#,##0.00'
    # Also set 基本预备费 item formula
    for e in entries:
        if e['seq'] == '18' and '基本预备费' in e['name']:
            ws.cell(row=e['row'], column=7).value = reserve_formula
            ws.cell(row=e['row'], column=6).value = reserve_formula
            ws.cell(row=e['row'], column=7).number_format = '#,##0.00'
            ws.cell(row=e['row'], column=6).number_format = '#,##0.00'

    # 建设期利息 section: = child item
    ws.cell(row=interest_row, column=7).value = f'=G{interest_row+1}'
    ws.cell(row=interest_row, column=6).value = f'=G{interest_row}'
    ws.cell(row=interest_row, column=7).number_format = '#,##0.00'
    ws.cell(row=interest_row, column=6).number_format = '#,##0.00'

    # Style section rows C-G
    for sec_row in [eng_row, other_row, reserve_row, interest_row]:
        for col in range(3, 8):
            c = ws.cell(row=sec_row, column=col)
            c.font = bold_font; c.alignment = right_align

    # ===== Grand Total formulas =====
    inv_row = None; work_row = None; total_row = None
    for e in entries:
        if e['name'] == '项目总投资':
            inv_row = e['row']
        if e.get('remark') and '按运营初期' in str(e.get('remark', '')):
            work_row = e['row']
        if e['name'] == '项目总资金' and e['row_type'] == 'grand_total':
            total_row = e['row']

    if inv_row:
        formula = f'=G{eng_row}+G{other_row}+G{reserve_row}+G{interest_row}'
        ws.cell(row=inv_row, column=7).value = formula
        ws.cell(row=inv_row, column=7).number_format = '#,##0.00'
        ws.cell(row=inv_row, column=7).font = bold_font
        ws.cell(row=inv_row, column=7).alignment = right_align

    if inv_row and work_row and total_row:
        formula = f'=G{inv_row}+G{work_row}'
        ws.cell(row=total_row, column=7).value = formula
        ws.cell(row=total_row, column=7).number_format = '#,##0.00'
        ws.cell(row=total_row, column=7).font = bold_font
        ws.cell(row=total_row, column=7).alignment = right_align

    # ===== Final formatting =====
    ws.freeze_panes = 'C6'
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # ========== Sheet 2: 编制说明 ==========
    ws2 = wb.create_sheet('编制说明')
    ws2.column_dimensions['A'].width = 4
    ws2.column_dimensions['B'].width = 88

    ws2.merge_cells('A1:B1')
    ws2['A1'] = '投资估算编制说明（修订版）'
    ws2['A1'].font = subtitle_font

    sections = [
        ('一、', '编制依据'),
        ('', '1. 本项目总平面图（黄许一期，2026.05.27 t8版）及建构筑物一览表——设计院施工图阶段设计文件'),
        ('', '2. 德阳市2025年上半年建设工程造价指数（183.16，以2005年为基期）'),
        ('', '3. 《四川省建设工程工程量清单计价定额》（2020版）'),
        ('', '4. 国家发展改革委《建设项目投资估算编审规程》（CECA/GC1）'),
        ('', '5. 《物资集散中心和陆港仓库建设规范》（T/SDTLA 0001-2024）'),
        ('', '6. 同类项目：德阳蓉欧冷链运营中心（已建成）、芜湖裕溪口多式联运基地（2025年招标）、德阳临港制造产业园（在建）'),
        ('', '7. 设备价格：2025年市场询价及同类项目采购参考'),
        ('', ''),
        ('二、', '主要取费标准'),
        ('', '· 建设单位管理费：财建〔2016〕504号，阶梯费率 → 530万元'),
        ('', '· 工程监理费：发改价格〔2007〕670号插入法 → 716万元'),
        ('', '· 勘察设计费：工程费用×2.5% → 1,193万元'),
        ('', '· 基本预备费：(工程费用+工程建设其他费)×8% → 4,554万元'),
        ('', '· 土地取得费：按180,000元/亩综合价（含出让金及税费），295亩 → 5,310万元'),
        ('', '· 建设期利息：按融资30,000万元×2.8%×2年（平均） → 1,680万元'),
        ('', '· 城市基础设施配套费：158,819㎡×20元/㎡ → 318万元'),
        ('', ''),
        ('三、', '主要单价指标来源'),
        ('', '· 2F钢结构物流仓 1,530~1,580元/㎡ — 增加二层楼面+货梯坡道土建+一级耐火，较单层方案+110~160元/㎡'),
        ('', '· 2F冷库（含保温+NH₃/CO₂复叠制冷）6,800元/㎡ — 规模13,440㎡，较旧版+300元/㎡'),
        ('', '· 2F中转仓（含双侧月台）1,620元/㎡'),
        ('', '· 消防工程（一级耐火）125元/㎡ — 一级耐火较二级+25元/㎡'),
        ('', '· 海绵城市增量 180万元 — 雨水花园+下凹绿地+透水铺装+调蓄池'),
        ('', '· 绿色建筑增量 120万元 — 一星级绿色建筑（GB/T 50378-2019）'),
        ('', '· 防火涂料增项 45元/㎡ — 厚涂型，钢结构一级耐火（钢柱≥3h）'),
        ('', ''),
        ('四、', '与旧版主要差异'),
        ('', '① 仓库由单层→双层钢结构门式刚架（1F→2F，建筑高度23.8m）'),
        ('', '② 耐火等级由二级→一级（消防工程单价100→125元/㎡，新增防火涂料715万元）'),
        ('', '③ 冷库面积由8,000→13,440㎡（+68%），设备规模相应增加'),
        ('', '④ 新增20台货梯（5t贯通式，500万元），充电桩48台（19万元）'),
        ('', '⑤ 新增海绵城市（180万元）+ 绿色建筑（120万元）+ BIM（127万元）'),
        ('', '⑥ 项目总资金由52,400→65,656万元，增幅约25.3%'),
        ('', ''),
        ('五、', '待确认事项'),
        ('', '① 土地出让金实际价格（暂按18万/亩综合价，需以自然资源部门评估价为准）'),
        ('', '② 冷库NH₃/CO₂复叠制冷系统的最终设备选型和价格（暂按国产+进口组合估算）'),
        ('', '③ 详勘后地基基础形式（如有软土层需增加桩基费用约500~800万元）'),
        ('', '④ 建设期利息按融资3亿元×2.8%×2年估算，实际利率以融资方案确定'),
        ('', '⑤ BIM应用深度（暂按基础建模+施工应用，若增加运维BIM需追加约80万元）'),
    ]

    for i, (k, v) in enumerate(sections):
        row = 3 + i
        c1 = ws2.cell(row=row, column=1, value=k)
        c1.font = bold_font if k else normal_font
        c2 = ws2.cell(row=row, column=2, value=v)
        c2.font = normal_font if k else normal_font

    # ========== Sheet 3: 投资构成分析 ==========
    ws3 = wb.create_sheet('投资构成分析')
    ws3.column_dimensions['A'].width = 32
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 14

    ws3.merge_cells('A1:C1')
    ws3['A1'] = '德阳物流基地（黄许一期）— 投资构成分析'
    ws3['A1'].font = subtitle_font

    for col, h in enumerate(['构成项目', '金额（万元）', '占比'], 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = Font(name='微软雅黑', bold=True, size=9)
        c.alignment = center_align; c.border = thin_border

    analysis = [
        # 工程费用 breakdown
        ('一、工程费用', f'=投资估算表!G{eng_row}', f'=投资估算表!G{eng_row}/投资估算表!G{total_row}', True),
        ('  1. 建筑工程', f'=投资估算表!G{sub_rows_eng[0]}',
         f'=投资估算表!G{sub_rows_eng[0]}/投资估算表!G{total_row}', False),
        ('  2. 安装工程', f'=投资估算表!G{sub_rows_eng[1]}',
         f'=投资估算表!G{sub_rows_eng[1]}/投资估算表!G{total_row}', False),
        ('  3. 设备购置费', f'=投资估算表!G{sub_rows_eng[2]}',
         f'=投资估算表!G{sub_rows_eng[2]}/投资估算表!G{total_row}', False),
        ('  4. 室外配套工程', f'=投资估算表!G{sub_rows_eng[3]}',
         f'=投资估算表!G{sub_rows_eng[3]}/投资估算表!G{total_row}', False),
        ('二、工程建设其他费', f'=投资估算表!G{other_row}',
         f'=投资估算表!G{other_row}/投资估算表!G{total_row}', True),
        ('三、预备费', f'=投资估算表!G{reserve_row}',
         f'=投资估算表!G{reserve_row}/投资估算表!G{total_row}', True),
        ('四、建设期利息', f'=投资估算表!G{interest_row}',
         f'=投资估算表!G{interest_row}/投资估算表!G{total_row}', True),
        ('五、铺底流动资金', f'=投资估算表!G{work_row}',
         f'=投资估算表!G{work_row}/投资估算表!G{total_row}', True),
        ('项目总资金', f'=投资估算表!G{total_row}', 1.0, True),
    ]

    for i, (name, amt_f, ratio, is_b) in enumerate(analysis):
        row = 4 + i
        ws3.cell(row=row, column=1, value=name).font = bold_font if is_b else normal_font
        ws3.cell(row=row, column=1).border = thin_border

        c = ws3.cell(row=row, column=2)
        if isinstance(amt_f, str) and amt_f.startswith('='):
            c.value = amt_f
        else:
            c.value = amt_f
        c.font = bold_font if is_b else normal_font
        c.number_format = '#,##0'
        c.alignment = right_align; c.border = thin_border

        c = ws3.cell(row=row, column=3)
        if isinstance(ratio, str):
            c.value = ratio
        else:
            c.value = ratio
        c.number_format = '0.0%'; c.alignment = center_align; c.border = thin_border

    # Also add key technical-economic indicators
    indicators_start = 4 + len(analysis) + 2
    ws3.merge_cells(f'A{indicators_start}:C{indicators_start}')
    ws3.cell(row=indicators_start, column=1, value='主要技术经济指标').font = subtitle_font

    indicators = [
        ('建设用地面积', '196,481 ㎡（约295亩）'),
        ('总建筑面积', '158,819 ㎡'),
        ('计容建筑面积', '317,411 ㎡'),
        ('容积率', '1.615'),
        ('建筑密度', '40.99%'),
        ('绿地率', '10.00%'),
        ('亩均投资', '222.5 万元/亩'),
        ('单位建筑面积综合投资', '4,134 元/㎡'),
    ]
    for i, (k, v) in enumerate(indicators):
        row = indicators_start + 1 + i
        ws3.cell(row=row, column=1, value=k).font = bold_font
        ws3.cell(row=row, column=1).border = thin_border
        c = ws3.cell(row=row, column=2, value=v)
        c.font = normal_font; c.border = thin_border
        ws3.merge_cells(f'B{row}:C{row}')

    # Save
    output_dir = output_dir or os.path.dirname(os.path.abspath(__file__))
    filename = os.path.join(output_dir, '..', 'assets', '德阳物流基地-投资估算表-修订版.xlsx')
    filename = os.path.abspath(filename)
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    wb.save(filename)
    print(f'Excel（含公式链接版·修订版）已生成: {filename}')
    print(f'项目总资金: 65,656万元')
    return filename

if __name__ == '__main__':
    generate_investment_table()
